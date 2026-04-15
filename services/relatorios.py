import pandas as pd
import re

from data.drive_loader import conectar_drive, listar_arquivos, baixar_arquivo
from openpyxl import Workbook
from openpyxl.styles import Font

# =========================
# 📥 EXPORTAÇÃO
# =========================
def exportar_excel(df: pd.DataFrame, nome_arquivo: str) -> str:
    df = df.copy()

    colunas_numericas = df.select_dtypes(include=["float", "int"]).columns
    for col in colunas_numericas:
        df[col] = df[col].round(2)

    caminho = f"{nome_arquivo}.xlsx"

    df.to_excel(caminho, index=False)

    print(f"Arquivo gerado: {caminho}")

    return caminho


# =========================
# 📊 NOVO: PARSE DO RELATÓRIO TXT
# =========================
def extrair_indicadores(texto):

    def buscar_valor(label):
        match = re.search(rf"{label}.*?:\s+([\d\.,-]+)", texto)
        if match:
            return float(match.group(1).replace(".", "").replace(",", "."))
        return 0

    dados = {
        "faturamento_bruto": buscar_valor("Faturamento Bruto"),
        "resultado_operacional": buscar_valor("RES. OPERACIONAL"),
        "sangria": buscar_valor("Sangria"),
        "troco": buscar_valor("Troco"),
    }

    # Ticket médio
    match_tm = re.search(r"TM-Ticket Medio por Cupom...:\s+([\d\.,]+)", texto)
    dados["ticket_medio"] = float(match_tm.group(1).replace(",", ".")) if match_tm else 0

    # Cupons
    match_tc = re.search(r"TC-Total Cupom..............:\s+(\d+)", texto)
    dados["cupons"] = int(match_tc.group(1)) if match_tc else 0

    return dados


# =========================
# 📊 RELATORIO POR DATA:
# =========================
def carregar_relatorio_por_data(data_input: str, folder_id: str):

    service = conectar_drive()
    arquivos = listar_arquivos(service, folder_id)

    if not arquivos:
        return None, "Nenhum arquivo encontrado no Drive."

    for arq in arquivos:
        if data_input in arq["name"]:
            texto = baixar_arquivo(service, arq["id"])
            return texto, arq["name"]

    return None, f"Nenhum relatório encontrado para a data {data_input}"


# ====================================
# 📊 EXTRAIR DADOS DO RELATORIO:
# ====================================
def extrair_resumo_relatorio(texto):

    def buscar_valor(label):
        match = re.search(rf"{label}.*?:\s+([\d\.,]+)", texto)
        return match.group(1) if match else "0,00"

    dados = {
        "faturamento_bruto": buscar_valor("Faturamento Bruto"),
        "tx_servico": buscar_valor("Tx. Serv Mesa"),
        "tc": buscar_valor("TC-Total Cupom"),
        "tm": buscar_valor("TM-Ticket Medio por Cupom"),
    }

    return dados


# ====================================
# 📊 EXTRAIR CANCELAMENTOS DO RELATORIO:
# ====================================
def extrair_cancelamentos(texto, tipo):

    if tipo == "antes":
        inicio = "CANCELAMENTO VENDA MESA (ANTES ENVIAR PRODUCAO)"
    else:
        inicio = "CANCELAMENTO VENDA MESA (DEPOIS ENVIAR PRODUCAO)"

    linhas = texto.split("\n")

    capturar = False
    dados = []

    for linha in linhas:

        if inicio in linha:
            capturar = True
            continue

        if capturar:
            if "TOTAL:" in linha:
                break

            if linha.strip() and linha.strip()[0].isdigit():
                dados.append(linha.strip())

    return dados


# ====================================
# 📊 FORMATAR COMO TABELA:
# ====================================
def formatar_tabela_cancelamentos(lista):

    if not lista:
        return "Sem registros"

    header = "Mesa | Produto | Qtde | Subtotal | Autorizado | Hora"
    linhas_formatadas = [header]

    for linha in lista:
        partes = linha.split()

        try:
            mesa = partes[0]
            produto = " ".join(partes[2:-4])
            qtd = partes[-4]
            subtotal = partes[-3]
            autor = partes[-2]
            hora = partes[-1]

            linha_formatada = f"{mesa} | {produto} | {qtd} | {subtotal} | {autor} | {hora}"
            linhas_formatadas.append(linha_formatada)

        except:
            continue

    return "\n".join(linhas_formatadas)


# ====================================
# 📊 EXTRAIR PRODUTOS RELATORIO:
# ====================================
def extrair_produtos_relatorio(texto):

    import pandas as pd

    linhas = texto.split("\n")

    capturar = False
    dados = []

    for linha in linhas:

        # 🔥 INÍCIO CORRETO
        if "MOVIMENTACAO DE PRODUTO" in linha:
            capturar = True
            continue

        if capturar:

            # 🔴 FIM DA SEÇÃO
            if "Periodo" in linha:
                break

            linha = linha.strip()

            # Ignorar linhas vazias ou separadores
            if not linha or linha.startswith("-"):
                continue

            # Captura linhas com dados (começam com código numérico)
            if linha[0].isdigit():

                partes = linha.split()

                try:
                    codigo = partes[0]

                    # Nome do produto (entre código e qtd)
                    nome = " ".join(partes[1:-3])

                    qtd = int(partes[-3])

                    valor = float(partes[-2].replace(",", "."))

                    dados.append({
                        "produto": nome,
                        "qtd": qtd,
                        "valor_total": valor
                    })

                except:
                    continue

    return pd.DataFrame(dados)


# ====================================
# 📊 FORMATAR COMO TABELA:
# ====================================
def gerar_excel_comissao(df, nome_arquivo="relatorio_comissao.xlsx"):

    wb = Workbook()
    ws = wb.active
    ws.title = "Comissão"

    bold = Font(bold=True)

    # Cabeçalho
    ws.cell(row=1, column=1, value="Data").font = bold

    for col_idx, col_name in enumerate(df.columns, start=2):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = bold

    # Dados
    for row_idx, (data, row) in enumerate(df.iterrows(), start=2):

        cell = ws.cell(row=row_idx, column=1, value=str(data))
        cell.font = bold

        for col_idx, value in enumerate(row, start=2):
            ws.cell(row=row_idx, column=col_idx, value=round(value, 2))

    # Linha TOTAL
    total_row = len(df) + 2
    ws.cell(row=total_row, column=1, value="TOTAL").font = bold

    for col_idx, col_name in enumerate(df.columns, start=2):
        total = df[col_name].sum()
        cell = ws.cell(row=total_row, column=col_idx, value=round(total, 2))
        cell.font = bold

    wb.save(nome_arquivo)

    return nome_arquivo


# ====================================
# EXTRAIR VENDAS POR HORA:
# ====================================
def extrair_vendas_por_hora(texto):

    linhas = texto.splitlines()

    dados = []

    for linha in linhas:

        # identifica linhas com padrão de horário
        if re.match(r"\d{2}:\d{2} - \d{2}:\d{2}", linha):

            partes = linha.split()

            try:
                periodo = partes[0]  # 11:00
                hora = int(periodo.split(":")[0])

                vendas = float(partes[3].replace(".", "").replace(",", "."))
                cupons = int(partes[4])
                ticket = float(partes[5].replace(",", "."))

                dados.append({
                    "hora": hora,
                    "faturamento": vendas,
                    "cupons": cupons,
                    "ticket": ticket
                })

            except:
                continue

    df = pd.DataFrame(dados)

    # 🔥 filtrar horário da operação
    df = df[(df["hora"] >= 11) & (df["hora"] <= 23)]

    return df

