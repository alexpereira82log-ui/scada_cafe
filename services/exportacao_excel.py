import os

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side,
)


def exportar_dataframe_excel(
    df,
    caminho,
):
    """
    Exporta um DataFrame para um arquivo Excel.

    Esta função será reutilizada por diversos
    módulos do sistema.
    """

    os.makedirs(
        os.path.dirname(caminho),
        exist_ok=True,
    )

    # ==========================================================
    # Exportação inicial
    # ==========================================================

    df.to_excel(
        caminho,
        index=False,
    )

    # ==========================================================
    # Abrir arquivo exportado
    # ==========================================================

    workbook = load_workbook(
        caminho
    )

    planilha = workbook.active

    # ==========================================================
    # Congelar cabeçalho
    # ==========================================================

    planilha.freeze_panes = "A2"

    # ==========================================================
    # Filtros
    # ==========================================================

    planilha.auto_filter.ref = planilha.dimensions

    # ==========================================================
    # Ajuste automático da largura das colunas
    # ==========================================================

    for coluna in planilha.columns:

        tamanho = max(
            len(str(celula.value or ""))
            for celula in coluna
        )

        letra = get_column_letter(
            coluna[0].column
        )

        planilha.column_dimensions[
            letra
        ].width = tamanho + 2

    # ==========================================================
    # Formatação do cabeçalho
    # ==========================================================

    cabecalho_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    cabecalho_font = Font(
        bold=True,
        color="FFFFFF",
    )

    cabecalho_alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    for celula in planilha[1]:

        celula.fill = cabecalho_fill
        celula.font = cabecalho_font
        celula.alignment = cabecalho_alignment

    # ==========================================================
    # Formatação dos dados
    # ==========================================================

    FORMATO_DATA = "dd/mm/yyyy"

    FORMATO_MOEDA = '#,##0.00'

    # ==========================================================
    # Formatação da coluna Data
    # ==========================================================

    for celula in planilha["A"][1:]:

        celula.number_format = FORMATO_DATA

    # ==========================================================
    # Formatação das colunas monetárias
    # ==========================================================

    for coluna in planilha.iter_cols(
        min_col=2,
    ):

        for celula in coluna[1:]:

            if isinstance(
                celula.value,
                (int, float),
            ):

                celula.number_format = FORMATO_MOEDA

                celula.alignment = Alignment(
                    horizontal="right",
                    vertical="center",
                )

    # ==========================================================
    # Destaque da linha TOTAL
    # ==========================================================

    linha_total = planilha.max_row

    fill_total = PatternFill(
        fill_type="solid",
        fgColor="D9EAD3",
    )

    font_total = Font(
        bold=True,
    )

    borda_superior = Border(
        top=Side(
            style="medium",
        )
    )

    for celula in planilha[linha_total]:

        celula.fill = fill_total

        celula.font = font_total

        celula.border = borda_superior

    # ==========================================================
    # Salvar alterações
    # ==========================================================

    workbook.save(
        caminho
    )

    workbook.close()